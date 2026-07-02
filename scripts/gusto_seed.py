"""
gusto_seed.py
-------------
Seeds the Gusto sandbox company with all Tenant 1 employees.

For each row in Users_tenant1.xlsx it:
  1. Creates the employee  (POST /v1/companies/{uuid}/employees)
     → if already exists, looks up the existing employee by email
  2. Creates a job          (POST /v1/employees/{uuid}/jobs)
     → if already exists, looks up the existing job
  3. Creates compensation   (POST /v1/jobs/{uuid}/compensations)
     → if already exists, marks as skipped (not failed)

Re-runnable: employees/jobs that already exist are recovered via lookup
rather than skipped entirely, so compensation is always attempted.

Results are written to gusto_seed_results.csv next to this script.

Usage:
    set GUSTO_ACCESS_TOKEN=<your token>
    set GUSTO_COMPANY_UUID=<your company uuid>
    python gusto_seed.py

Optional:
    set GUSTO_API_BASE=https://api.gusto-demo.com   (default)
"""

import csv
import os
import sys
import time
from pathlib import Path
from typing import Optional

import httpx
import openpyxl

from dotenv import load_dotenv
load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────
ACCESS_TOKEN = os.environ.get("GUSTO_ACCESS_TOKEN")
COMPANY_UUID = os.environ.get("GUSTO_COMPANY_UUID")
API_BASE     = os.environ.get("GUSTO_API_BASE", "https://api.gusto-demo.com")

if not ACCESS_TOKEN or not COMPANY_UUID:
    sys.exit("ERROR: GUSTO_ACCESS_TOKEN and GUSTO_COMPANY_UUID must be set.")

HEADERS = {
    "Authorization":       f"Bearer {ACCESS_TOKEN}",
    "Content-Type":        "application/json",
    "X-Gusto-API-Version": "2026-06-15",
}

HIRE_DATE      = "2023-01-01"
EFFECTIVE_DATE = "2026-06-19"   # Gusto minimum: must not be in the past
PAYMENT_UNIT   = "Year"
PACE           = 0.3            # seconds between requests

SCRIPT_DIR  = Path(__file__).parent
INPUT_FILE  = SCRIPT_DIR / "Users_tenant1.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "gusto_seed_results.csv"

# ── Load spreadsheet ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active
col_names = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"Loaded {len(rows)} employees from {INPUT_FILE.name}")


# ── HTTP helpers ──────────────────────────────────────────────────────────────

def post(client: httpx.Client, url: str, payload: dict) -> tuple[int, dict]:
    resp = client.post(url, json=payload, headers=HEADERS)
    if resp.status_code >= 400:
        print(f"  DEBUG {resp.status_code} {resp.request.url}")
        print(f"  DEBUG body: {resp.text[:500]}")
    try:
        body = resp.json()
    except Exception:
        body = {"raw": resp.text}
    return resp.status_code, body


def get(client: httpx.Client, url: str, params: dict | None = None) -> tuple[int, dict | list]:
    resp = client.get(url, headers=HEADERS, params=params or {})
    try:
        body = resp.json()
    except Exception:
        body = {"raw": resp.text}
    return resp.status_code, body


def handle_rate_limit(client: httpx.Client, url: str, payload: dict) -> tuple[int, dict]:
    status, body = post(client, url, payload)
    if status == 429:
        print("  Rate limited — sleeping 60s and retrying")
        time.sleep(60)
        status, body = post(client, url, payload)
    return status, body


# ── Lookup helpers (for re-run resilience) ────────────────────────────────────

def find_existing_employee(client: httpx.Client, email: str) -> Optional[str]:
    """Return the Gusto employee UUID matching the given email, or None."""
    status, body = get(
        client,
        f"{API_BASE}/v1/companies/{COMPANY_UUID}/employees",
        params={"include": "all_compensations"},
    )
    time.sleep(PACE)
    if status != 200 or not isinstance(body, list):
        return None
    email_lower = email.lower()
    for emp in body:
        if email_lower in (
            (emp.get("work_email") or "").lower(),
            (emp.get("email")      or "").lower(),
        ):
            return emp["uuid"]
    return None


def find_existing_job(client: httpx.Client, employee_uuid: str) -> Optional[str]:
    """Return the first job UUID for the given employee, or None."""
    status, body = get(client, f"{API_BASE}/v1/employees/{employee_uuid}/jobs")
    time.sleep(PACE)
    if status != 200 or not isinstance(body, list) or not body:
        return None
    return body[0]["uuid"]


# ── Main loop ─────────────────────────────────────────────────────────────────
results = []
created = skipped = failed = 0

with httpx.Client(timeout=30) as client:
    for i, row in enumerate(rows, start=1):
        r      = dict(zip(col_names, row))
        first  = r["FirstName"]
        last   = r["LastName"]
        email  = r["userPrincipalName"]
        title  = r["Title"] or "Employee"
        salary = float(r["Salary"])

        print(f"[{i:3d}/{len(rows)}] {email}")

        result = {
            "email":             email,
            "status":            "",
            "employee_uuid":     "",
            "job_uuid":          "",
            "compensation_uuid": "",
            "error":             "",
        }

        # ── Step 1: Create employee ───────────────────────────────────────────
        status, body = handle_rate_limit(
            client,
            f"{API_BASE}/v1/companies/{COMPANY_UUID}/employees",
            {
                "first_name":    first,
                "last_name":     last,
                "email":         email,
                "date_of_birth": "1980-01-01",  # required by Gusto; placeholder
            },
        )
        time.sleep(PACE)

        if status == 201:
            employee_uuid = body["uuid"]
            result["employee_uuid"] = employee_uuid
            print(f"         employee created  {employee_uuid}")
        elif status == 422:
            # Already exists — look up the UUID so we can continue
            print(f"         employee exists — looking up UUID")
            employee_uuid = find_existing_employee(client, email)
            if not employee_uuid:
                result["status"] = "failed"
                result["error"]  = "employee 422 but could not find existing UUID"
                failed += 1
                print(f"         FAILED — could not find existing employee")
                results.append(result)
                continue
            result["employee_uuid"] = employee_uuid
            print(f"         employee found    {employee_uuid}")
        else:
            result["status"] = "failed"
            result["error"]  = f"employee {status}: {str(body)[:200]}"
            failed += 1
            print(f"         FAILED employee HTTP {status}")
            results.append(result)
            continue

        # ── Step 2: Create job ────────────────────────────────────────────────
        status, body = handle_rate_limit(
            client,
            f"{API_BASE}/v1/employees/{employee_uuid}/jobs",
            {
                "title":     title,
                "hire_date": HIRE_DATE,
                "primary":   True,
            },
        )
        time.sleep(PACE)

        if status == 201:
            job_uuid = body["uuid"]
            result["job_uuid"] = job_uuid
            print(f"         job created        {job_uuid}")
        elif status == 422:
            # Already exists — look up the UUID so we can continue
            print(f"         job exists — looking up UUID")
            job_uuid = find_existing_job(client, employee_uuid)
            if not job_uuid:
                result["status"] = "failed"
                result["error"]  = "job 422 but could not find existing UUID"
                failed += 1
                print(f"         FAILED — could not find existing job")
                results.append(result)
                continue
            result["job_uuid"] = job_uuid
            print(f"         job found          {job_uuid}")
        else:
            result["status"] = "failed"
            result["error"]  = f"job {status}: {str(body)[:200]}"
            failed += 1
            print(f"         FAILED job HTTP {status}")
            results.append(result)
            continue

        # ── Step 3: Create compensation ───────────────────────────────────────
        status, body = handle_rate_limit(
            client,
            f"{API_BASE}/v1/jobs/{job_uuid}/compensations",
            {
                "rate":           str(salary),
                "payment_unit":   PAYMENT_UNIT,
                "effective_date": EFFECTIVE_DATE,
                "flsa_status":    "Exempt",
            },
        )
        time.sleep(PACE)

        if status == 201:
            result["compensation_uuid"] = body["uuid"]
            result["status"] = "created"
            created += 1
            print(f"         compensation set   {salary:,.0f}/yr")
        elif status == 422:
            # Compensation already exists — not a failure
            result["status"] = "skipped"
            result["error"]  = "compensation already exists"
            skipped += 1
            print(f"         compensation already exists — skipped")
        else:
            result["status"] = "failed"
            result["error"]  = f"compensation {status}: {str(body)[:200]}"
            failed += 1
            print(f"         FAILED compensation HTTP {status}")

        results.append(result)

# ── Write results ─────────────────────────────────────────────────────────────
with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=["email", "status", "employee_uuid", "job_uuid", "compensation_uuid", "error"],
    )
    writer.writeheader()
    writer.writerows(results)

print(f"\nDone.  Created={created}  Skipped={skipped}  Failed={failed}")
print(f"Results → {OUTPUT_FILE}")
