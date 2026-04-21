"""
export_utils.py
---------------
Excel export helpers for the Award Nomination analytics.

Public API
----------
build_finding_workbook(data: dict) -> io.BytesIO
    Accepts the dict returned by sqlhelper2.get_finding_with_nominations()
    and returns a ready-to-stream BytesIO containing the .xlsx file.
"""

import io
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Pattern display metadata (mirrors PATTERN_META on the frontend) ───────────

PATTERN_LABEL = {
    "Ring":                  "Nomination Ring",
    "SuperNominator":        "Super Nominator",
    "Desert":                "Nomination Desert",
    "ApproverAffinity":      "Approver Affinity",
    "CopyPaste":             "Copy-Paste Fraud",
    "TransactionalLanguage": "Transactional Language",
    "HiddenCandidate":       "Hidden Candidate",
}

PATTERN_DESC = {
    "Ring":                  "Directed cycle of mutual nominations",
    "SuperNominator":        "Unusually high nomination volume",
    "Desert":                "Entire team absent from all nominations",
    "ApproverAffinity":      "Elevated approval rate for specific pair",
    "CopyPaste":             "Near-identical nomination descriptions",
    "TransactionalLanguage": "Personal-benefit phrasing in description",
    "HiddenCandidate":       "Named in descriptions but never nominated",
}

# Three colors — sufficient to color any cycle (graph map theorem)
_USER_PALETTE = [
    "BDD7EE",  # 🔵 blue
    "FFEB9C",  # 🟡 yellow
    "C6EFCE",  # 🟢 green
]


# ── Nomination ordering ───────────────────────────────────────────────────────

def _parse_ring_start_user(detail: str) -> int | None:
    """
    Extract the first user ID from the ring detail string.
    Expected format: "Nomination ring of N users: 3 → 45 → … → 3 (total…)"
    Returns the integer ID, or None if parsing fails.
    """
    m = re.search(r":\s*(\d+)\s*→", detail or "")
    return int(m.group(1)) if m else None


def _order_nominations(pattern: str, nominations: list[dict], detail: str = "") -> list[dict]:
    """
    For Ring findings: group nominations by directed edge (nominator→beneficiary),
    walk the unique edges in cycle order starting from the first user named in
    the Detail string, and emit all nominations for each edge together (sorted
    by date).  Multiple nominations between the same pair are kept consecutively.

    For all other patterns: sort by nomination date.
    """
    if pattern != "Ring" or not nominations:
        return sorted(nominations, key=lambda n: n["nominationDate"])

    edge_groups: dict[tuple, list] = defaultdict(list)
    for n in nominations:
        edge_groups[(n["nominatorId"], n["beneficiaryId"])].append(n)
    for group in edge_groups.values():
        group.sort(key=lambda x: x["nominationDate"])

    unique_edges      = list(edge_groups.keys())
    edge_by_nominator = {edge[0]: edge for edge in unique_edges}

    # Anchor the walk at the first user mentioned in the detail text so the
    # spreadsheet order matches the "A → B → C → … → A" description exactly.
    anchor_uid = _parse_ring_start_user(detail)
    start_edge = edge_by_nominator.get(anchor_uid, unique_edges[0])

    ordered_edges = [start_edge]
    current_edge  = start_edge
    for _ in range(len(unique_edges) - 1):
        nxt = edge_by_nominator.get(current_edge[1])
        if not nxt or nxt == start_edge:
            break
        ordered_edges.append(nxt)
        current_edge = nxt

    if len(ordered_edges) == len(unique_edges):
        return [n for edge in ordered_edges for n in edge_groups[edge]]

    # Fallback if chain is broken
    return sorted(nominations, key=lambda n: (n["nominatorId"], n["nominationDate"]))


# ── Public entry point ────────────────────────────────────────────────────────

def build_finding_workbook(data: dict) -> io.BytesIO:
    """
    Build an Excel workbook for a single integrity finding.

    Parameters
    ----------
    data : dict
        As returned by sqlhelper2.get_finding_with_nominations().

    Returns
    -------
    io.BytesIO
        Seeked to position 0, ready to be streamed.
    """
    finding_id  = data["finding_id"] if "finding_id" in data else data["findingId"]
    pattern     = data["patternType"]
    nominations = _order_nominations(pattern, data["nominations"], data.get("detail", ""))

    # ── Per-user color assignment ─────────────────────────────────────────────
    user_color_index: dict[int, int] = {}

    def color_for(uid: int | None) -> str | None:
        if uid is None:
            return None
        if uid not in user_color_index:
            user_color_index[uid] = len(user_color_index) % len(_USER_PALETTE)
        return _USER_PALETTE[user_color_index[uid]]

    def user_fill(uid: int | None) -> PatternFill | None:
        hex_color = color_for(uid)
        return PatternFill("solid", start_color=hex_color, end_color=hex_color) if hex_color else None

    def fmt_name(name: str, uid: int | None) -> str:
        return f"{name} ({uid})" if uid is not None else name

    # Pre-scan in display order so colors are stable across all columns
    for nom in nominations:
        color_for(nom["nominatorId"])
        color_for(nom["beneficiaryId"])
        color_for(nom.get("approverId"))

    # ── Styles ────────────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", start_color="2F5496", end_color="2F5496")
    META_FILL   = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    NEUTRAL_ALT = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
    LABEL_FONT  = Font(name="Arial", bold=True, size=10)
    VALUE_FONT  = Font(name="Arial", size=10)
    HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    THIN        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finding {finding_id}"

    def meta_row(row: int, label: str, value) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = LABEL_FONT
        lc.fill = META_FILL
        lc.border = BORDER
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = VALUE_FONT
        vc.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)

    meta_row(1, "Finding type:",        PATTERN_LABEL.get(pattern, pattern))
    meta_row(2, "Finding #",            finding_id)
    meta_row(3, "Finding Desc.:",       PATTERN_DESC.get(pattern, ""))
    meta_row(4, "Finding explanation:", data["detail"])
    meta_row(5, "Total approved/Paid:", data["totalAmount"])
    ws.cell(row=5, column=2).number_format = "#,##0.00"

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        "Nomination Id", "Nominator", "Beneficiary", "Approver",
        "Amount", "Currency", "Description", "Status", "Date",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border    = BORDER

    # ── Nomination rows ───────────────────────────────────────────────────────
    for i, nom in enumerate(nominations):
        r       = i + 7
        neutral = NEUTRAL_ALT if i % 2 == 1 else None

        row_data = [
            (nom["nominationId"],                                          neutral),
            (fmt_name(nom["nominatorName"],   nom["nominatorId"]),         user_fill(nom["nominatorId"])),
            (fmt_name(nom["beneficiaryName"], nom["beneficiaryId"]),       user_fill(nom["beneficiaryId"])),
            (fmt_name(nom["approverName"],    nom.get("approverId")),      user_fill(nom.get("approverId"))),
            (nom["amount"],                                                neutral),
            (nom["currency"],                                              neutral),
            (nom["description"],                                           neutral),
            (nom["status"],                                                neutral),
            (nom["nominationDate"],                                        neutral),
        ]
        for col, (val, fill) in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font   = VALUE_FONT
            c.border = BORDER
            if fill:
                c.fill = fill
        ws.cell(row=r, column=5).number_format = "#,##0.00"

    # ── Column widths & freeze ────────────────────────────────────────────────
    for col, w in enumerate([14, 26, 26, 26, 12, 10, 50, 14, 13], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
