#!/usr/bin/env python3
"""
Analytics Export MCP Server
Exports analytics data to Excel, PDF, and CSV formats via MCP tools.
"""

import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from dataclasses import dataclass

from azure_blob import upload_export

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import pandas as pd

# MCP imports
from mcp.server import Server
from mcp.types import Tool, TextContent

# ============================================================================
# SETUP: Logging
# ============================================================================

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================

EXPORT_BASE_PATH = Path(os.getenv("EXPORT_BASE_PATH", "./exports"))
EXPORT_ARCHIVE_DAYS = int(os.getenv("EXPORT_ARCHIVE_DAYS", "7"))
EXPORT_ACTIVE_DIR = EXPORT_BASE_PATH / "active"
EXPORT_ARCHIVE_DIR = EXPORT_BASE_PATH / "archive"

# PDF styling
PDF_PAGE_WIDTH = float(os.getenv("PDF_PAGE_WIDTH", "210"))  # mm
PDF_PAGE_HEIGHT = float(os.getenv("PDF_PAGE_HEIGHT", "297"))
PDF_FONT_NAME = os.getenv("PDF_FONT_NAME", "Helvetica")
PDF_FONT_SIZE = int(os.getenv("PDF_FONT_SIZE", "11"))

# Excel styling
EXCEL_HEADER_COLOR = os.getenv("EXCEL_HEADER_COLOR", "4472C4")  # Blue
EXCEL_AUTO_FIT = os.getenv("EXCEL_AUTO_FIT", "true").lower() == "true"

# Create directories
EXPORT_ACTIVE_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def _get_timestamp_str() -> str:
    """Get formatted timestamp for filenames."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _sanitize_filename(filename: str) -> str:
    """Remove/replace invalid filename characters."""
    return "".join(c if c.isalnum() or c in "._- " else "_" for c in filename)

def _cleanup_old_files():
    """Move files older than EXPORT_ARCHIVE_DAYS to archive."""
    cutoff_date = datetime.now() - timedelta(days=EXPORT_ARCHIVE_DAYS)
    
    for file_path in EXPORT_ACTIVE_DIR.glob("*"):
        if file_path.is_file():
            mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
            if mtime < cutoff_date:
                archive_path = EXPORT_ARCHIVE_DIR / file_path.name
                file_path.rename(archive_path)
                logger.info(f"Archived: {file_path.name}")

def _parse_data_table(data_table: Optional[list]) -> pd.DataFrame:
    """Convert list of dicts to pandas DataFrame."""
    if not data_table:
        return None
    try:
        return pd.DataFrame(data_table)
    except Exception as e:
        logger.warning(f"Failed to parse data_table: {e}")
        return None

# ============================================================================
# EXCEL EXPORT
# ============================================================================

def export_to_excel(question: str, answer: str, data_table: Optional[list] = None, filename: Optional[str] = None) -> dict:
    """
    Export analytics to Excel workbook.
    
    Args:
        question: User's question
        answer: LLM response
        data_table: Optional list of dicts (rows)
        filename: Optional custom filename
        
    Returns:
        Dict with status, file_path, file_size_bytes, rows_exported
    """
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Styling
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, 
                                   end_color=EXCEL_HEADER_COLOR, 
                                   fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary sheet
        title_cell = ws["A1"]
        title_cell.value = "Analytics Report"
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        
        # Timestamp
        ts_cell = ws["A2"]
        ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ts_cell.font = Font(italic=True, size=10)
        
        # Question
        ws["A4"].value = "Question:"
        ws["A4"].font = Font(bold=True)
        ws.merge_cells("B4:D4")
        ws["B4"].value = question
        
        # Answer
        ws["A6"].value = "Answer:"
        ws["A6"].font = Font(bold=True)
        ws.merge_cells("A7:D15")
        answer_cell = ws["A7"]
        answer_cell.value = answer
        answer_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        
        # Data sheet (if provided)
        if data_table:
            df = _parse_data_table(data_table)
            if df is not None:
                ws_data = wb.create_sheet("Data")
                
                # Headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws_data.cell(row=1, column=col_idx)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for row_idx, row_data in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_data.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border
                
                # Auto-fit columns
                if EXCEL_AUTO_FIT:
                    for col in ws_data.columns:
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws_data.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Save file
        if filename is None:
            filename = f"analytics_export_{_get_timestamp_str()}_xlsx.xlsx"
        else:
            filename = _sanitize_filename(filename)
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
        
        file_path = EXPORT_ACTIVE_DIR / filename
        wb.save(file_path)
        
        file_size = file_path.stat().st_size
        rows_exported = len(data_table) if data_table else 0
        
        logger.info(f"Excel export: {file_path} ({file_size} bytes, {rows_exported} rows)")

        download_url = upload_export(file_path)
        
        return {
            "status": "success",
            "download_url": download_url,
            "file_size_bytes": file_size,
            "rows_exported": rows_exported
        }
    
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return {
            "status": "error",
            "error_code": "IO_ERROR",
            "message": str(e)
        }

# ============================================================================
# PDF EXPORT
# ============================================================================

def export_to_pdf(question: str, answer: str, data_table: Optional[list] = None, 
                  filename: Optional[str] = None, include_timestamp: bool = True) -> dict:
    """
    Export analytics to PDF report.
    
    Args:
        question: User's question
        answer: LLM response
        data_table: Optional list of dicts (rows)
        filename: Optional custom filename
        include_timestamp: Whether to include generation timestamp
        
    Returns:
        Dict with status, file_path, file_size_bytes, pages
    """
    try:
        # Generate filename
        if filename is None:
            filename = f"analytics_export_{_get_timestamp_str()}_pdf.pdf"
        else:
            filename = _sanitize_filename(filename)
            if not filename.endswith(".pdf"):
                filename += ".pdf"
        
        file_path = EXPORT_ACTIVE_DIR / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=15*mm,
            rightMargin=15*mm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#4472C4"),
            spaceAfter=10,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor("#4472C4"),
            spaceAfter=6,
            spaceBefore=6
        )
        
        # Story (content)
        story = []
        
        # Title
        story.append(Paragraph("Analytics Report", title_style))
        
        # Timestamp (if requested)
        if include_timestamp:
            ts_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(ts_text, styles['Normal']))
        
        story.append(Spacer(1, 5*mm))
        
        # Question section
        story.append(Paragraph("Question", heading_style))
        story.append(Paragraph(question, styles['BodyText']))
        story.append(Spacer(1, 5*mm))
        
        # Answer section
        story.append(Paragraph("Response", heading_style))
        # Wrap answer text
        answer_lines = answer.split('\n')
        for line in answer_lines:
            if line.strip():
                story.append(Paragraph(line, styles['BodyText']))
        
        # Data table section (if provided)
        page_count = 1
        if data_table:
            df = _parse_data_table(data_table)
            if df is not None and len(df) > 0:
                story.append(Spacer(1, 5*mm))
                story.append(PageBreak())
                story.append(Paragraph("Data Export", heading_style))
                story.append(Spacer(1, 3*mm))
                
                # Convert DataFrame to table data
                table_data = [list(df.columns)] + df.values.tolist()
                
                # Limit rows per page
                rows_per_page = 20
                for i in range(0, len(table_data), rows_per_page + 1):
                    chunk = table_data[i:i + rows_per_page + 1]
                    
                    # Create table
                    table = Table(chunk, colWidths=[60*mm / len(df.columns)] * len(df.columns))
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ]))
                    
                    story.append(table)
                    if i + rows_per_page + 1 < len(table_data):
                        story.append(Spacer(1, 3*mm))
                        story.append(PageBreak())
                        page_count += 1
        
        # Build PDF
        doc.build(story)
        
        file_size = file_path.stat().st_size
        
        logger.info(f"PDF export: {file_path} ({file_size} bytes, {page_count} pages)")

        download_url = upload_export(file_path)
        
        return {
            "status": "success",
            "download_url": download_url,
            "file_size_bytes": file_size,
            "pages": page_count
        }
    
    except Exception as e:
        logger.error(f"PDF export failed: {e}")
        return {
            "status": "error",
            "error_code": "IO_ERROR",
            "message": str(e)
        }

# ============================================================================
# CSV EXPORT
# ============================================================================

def export_to_csv(data_table: list, filename: Optional[str] = None) -> dict:
    """
    Export data table to CSV.
    
    Args:
        data_table: List of dicts (required)
        filename: Optional custom filename
        
    Returns:
        Dict with status, file_path, file_size_bytes, rows_exported
    """
    try:
        if not data_table:
            return {
                "status": "error",
                "error_code": "INVALID_DATA",
                "message": "data_table is required for CSV export"
            }
        
        # Parse to DataFrame
        df = _parse_data_table(data_table)
        if df is None:
            return {
                "status": "error",
                "error_code": "INVALID_DATA",
                "message": "Failed to parse data_table"
            }
        
        # Generate filename
        if filename is None:
            filename = f"analytics_export_{_get_timestamp_str()}_csv.csv"
        else:
            filename = _sanitize_filename(filename)
            if not filename.endswith(".csv"):
                filename += ".csv"
        
        file_path = EXPORT_ACTIVE_DIR / filename
        df.to_csv(file_path, index=False)
        
        file_size = file_path.stat().st_size
        rows_exported = len(df)
        
        logger.info(f"CSV export: {file_path} ({file_size} bytes, {rows_exported} rows)")

        download_url = upload_export(file_path)
        
        return {
            "status": "success",
            "download_url": download_url,
            "file_size_bytes": file_size,
            "rows_exported": rows_exported
        }
    
    except Exception as e:
        logger.error(f"CSV export failed: {e}")
        return {
            "status": "error",
            "error_code": "IO_ERROR",
            "message": str(e)
        }

# ============================================================================
# LIST EXPORTED FILES
# ============================================================================

def list_exported_files(limit: int = 20, format_filter: Optional[str] = None) -> dict:
    """
    List exported files.
    
    Args:
        limit: Max results
        format_filter: Filter by 'xlsx', 'pdf', 'csv', or None
        
    Returns:
        Dict with status, files list, total count
    """
    try:
        files = []
        
        for file_path in sorted(EXPORT_ACTIVE_DIR.glob("*"), key=lambda p: p.stat().st_mtime, reverse=True):
            if file_path.is_file():
                # Check format filter
                suffix = file_path.suffix.lstrip(".")
                if format_filter and suffix != format_filter:
                    continue
                
                files.append({
                    "filename": file_path.name,
                    "format": suffix,
                    "created_at": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                    "size_bytes": file_path.stat().st_size,
                    "file_path": str(file_path)
                })
        
        # Apply limit
        files = files[:limit]
        
        return {
            "status": "success",
            "files": files,
            "total": len(files)
        }
    
    except Exception as e:
        logger.error(f"List files failed: {e}")
        return {
            "status": "error",
            "error_code": "IO_ERROR",
            "message": str(e)
        }

# ============================================================================
# MCP SERVER
# ============================================================================

server = Server("analytics-export-service")

@server.list_tools()
async def handle_list_tools():
    """List available tools."""
    return [
        Tool(
            name="export_to_excel",
            description="Export analytics data to Excel workbook with formatting",
            inputSchema={
                "type": "object",
                "properties": {
                    "question": {"type": "string", "description": "User's analytics question"},
                    "answer": {"type": "string", "description": "LLM-generated response"},
                    "data_table": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "Optional: structured data rows (list of dicts)"
                    },
                    "filename": {"type": "string", "description": "Optional: custom filename"}
                },
                "required": ["question", "answer"]
            }
        ),
        Tool(
            name="export_to_pdf",
            description="Export analytics data to PDF report",
            inputSchema={
                "type": "object",
                "properties": {
                    "question": {"type": "string", "description": "User's analytics question"},
                    "answer": {"type": "string", "description": "LLM-generated response"},
                    "data_table": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "Optional: structured data rows (list of dicts)"
                    },
                    "filename": {"type": "string", "description": "Optional: custom filename"},
                    "include_timestamp": {"type": "boolean", "description": "Include generation timestamp (default: true)"}
                },
                "required": ["question", "answer"]
            }
        ),
        Tool(
            name="export_to_csv",
            description="Export data table to CSV format",
            inputSchema={
                "type": "object",
                "properties": {
                    "data_table": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "Structured data rows (list of dicts, required)"
                    },
                    "filename": {"type": "string", "description": "Optional: custom filename"}
                },
                "required": ["data_table"]
            }
        ),
        Tool(
            name="list_exported_files",
            description="List previously generated export files",
            inputSchema={
                "type": "object",
                "properties": {
                    "limit": {"type": "integer", "description": "Max results (default: 20)"},
                    "format_filter": {"type": "string", "description": "Filter by 'xlsx', 'pdf', 'csv'"}
                }
            }
        )
    ]

@server.call_tool()
async def handle_call_tool(name: str, arguments: dict[str, Any]):
    """Handle tool calls."""
    _cleanup_old_files()
    
    if name == "export_to_excel":
        result = export_to_excel(
            question=str(arguments.get("question")),
            answer=str(arguments.get("answer")),
            data_table=arguments.get("data_table"),
            filename=arguments.get("filename")
        )
    elif name == "export_to_pdf":
        result = export_to_pdf(
            question=arguments.get("question"),
            answer=arguments.get("answer"),
            data_table=arguments.get("data_table"),
            filename=arguments.get("filename"),
            include_timestamp=arguments.get("include_timestamp", True)
        )
    elif name == "export_to_csv":
        result = export_to_csv(
            data_table=arguments.get("data_table"),
            filename=arguments.get("filename")
        )
    elif name == "list_exported_files":
        result = list_exported_files(
            limit=arguments.get("limit", 20),
            format_filter=arguments.get("format_filter")
        )
    else:
        result = {"status": "error", "error_code": "UNKNOWN_TOOL", "message": f"Unknown tool: {name}"}
    
    return [TextContent(type="text", text=json.dumps(result, indent=2))]

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import asyncio
    import sys
    from mcp.server.stdio import stdio_server
    
    logger.info("Starting Analytics Export MCP Server...")
    logger.info(f"Export directory: {EXPORT_ACTIVE_DIR}")
    
    # Run the server using stdio transport
    async def main():
        async with stdio_server() as (read_stream, write_stream):
            logger.info("Server started and ready for requests")
            await server.run(read_stream, write_stream, server.create_initialization_options())
    
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Server stopped")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Server error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)